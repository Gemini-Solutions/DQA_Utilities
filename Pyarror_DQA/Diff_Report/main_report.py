#!/usr/bin/env python3

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import date, timedelta

# Sample data generation
num_securities = 1000
tickers = [f"TICKER{i}" for i in range(1, num_securities + 1)]

# Generate random dates for issue and maturity
start_date = date(2000, 1, 1)
issue_dates = [start_date + timedelta(days=np.random.randint(0, 365*20)) for _ in range(num_securities)]
maturity_dates = [issue_date + timedelta(days=np.random.randint(365*5, 365*30)) for issue_date in issue_dates]

# Create additional columns for Reference and Market Data
columns = ["Issuer Name", "Issue Date", "Maturity Date", "Coupon Rate (%)", "Credit Rating", "Interest Rate", 
           "Price", "Yield (%)", "Volume", "Bid Price", "Ask Price"]

# Define random data for the two sources
data1 = pd.DataFrame({
    "Ticker": tickers,
    "Issuer Name": [f"Issuer {i}" for i in range(1, num_securities + 1)],
    "Issue Date": issue_dates,
    "Maturity Date": maturity_dates,
    "Coupon Rate (%)": np.random.uniform(0, 10, num_securities),
    "Credit Rating": np.random.choice(["AAA", "AA", "A", "BBB", "BB", "B", "CCC"], num_securities),
    "Interest Rate": np.random.uniform(0, 5, num_securities),
    "Price": np.random.uniform(80, 120, num_securities),
    "Yield (%)": np.random.uniform(0.5, 8, num_securities),
    "Volume": np.random.randint(1000, 100000, num_securities),
    "Bid Price": np.random.uniform(80, 120, num_securities),
    "Ask Price": np.random.uniform(80, 120, num_securities)
})
data2 = data1.copy()

# Introduce random differences in data2
for _ in range(150):  # Change 150 random entries
    idx = np.random.choice(data2.index)
    col = np.random.choice(columns)

    # Check if the column has numerical data
    if pd.api.types.is_numeric_dtype(data2[col]):
        # Add a random float to numeric columns
        modified_value = data1.at[idx, col] + np.random.normal(0, 5)

        # Cast to int if the original column type is integer
        if pd.api.types.is_integer_dtype(data2[col]):
            modified_value = int(modified_value)

        data2.at[idx, col] = modified_value
    else:
        # For string-based columns, change the value to a new random option
        if col == "Issuer Name":
            data2.at[idx, col] = f"Issuer {np.random.randint(1, num_securities + 1)}"
        elif col == "Credit Rating":
            data2.at[idx, col] = np.random.choice(["AAA", "AA", "A", "BBB", "BB", "B", "CCC"])


for _ in range(100):  # Remove data in 100 random entries in data1 and data2
    idx = np.random.choice(data1.index)
    col = np.random.choice(columns[1:])
    if np.random.rand() > 0.5:
        data1.at[idx, col] = np.nan
    else:
        data2.at[idx, col] = np.nan

# Reformat `comparison_df` to have left and right values in consecutive columns
comparison_df = pd.DataFrame({"Ticker": data1["Ticker"]})

for col in columns:
    comparison_df[f"{col} (Left)"] = data1[col]
    comparison_df[f"{col} (Right)"] = data2[col]

# Compute summary statistics
no_diff_count = 0
data_diff_count = 0
only_left_count = 0
only_right_count = 0

rows_with_differences = []
rows_with_no_differences = []

for idx, row in comparison_df.iterrows():
    has_diff = False
    only_left = False
    only_right = False

    for i in range(1, len(columns) * 2, 2):  # Iterate through left-right pairs
        left_val = row.iloc[i]
        right_val = row.iloc[i + 1]

        if pd.isna(left_val) and not pd.isna(right_val):
            only_right = True
        elif not pd.isna(left_val) and pd.isna(right_val):
            only_left = True
        elif left_val != right_val:
            has_diff = True
            only_left = only_right = False

    if has_diff or only_left or only_right:
        rows_with_differences.append(row)
    else:
        rows_with_no_differences.append(row)

# Calculate counts for the summary DataFrame
data_diff_count = len(rows_with_differences)
only_left_count = sum(1 for row in rows_with_differences if row.iloc[1::2].isna().any())
only_right_count = sum(1 for row in rows_with_differences if row.iloc[::2].isna().any())
no_diff_count = len(rows_with_no_differences)

# Prepare summary DataFrame
summary_df = pd.DataFrame({
    "Description": [
        "Rows with no differences",
        "Rows with data differences",
        "Rows with only data in left source",
        "Rows with only data in right source"
    ],
    "Count": [
        no_diff_count,
        data_diff_count,
        only_left_count,
        only_right_count
    ],
    "Color": ["", "Orange", "Yellow", "Pink"]
})

# Define fills for coloring
light_green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Light Green
thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

# Write to Excel with conditional formatting
wb = Workbook()
summary_sheet = wb.active
summary_sheet.title = "Summary & Legend"

# Populate Summary Sheet
for r in dataframe_to_rows(summary_df, index=False, header=True):
    summary_sheet.append(r)

# Apply styling to header row in Summary Sheet
for cell in summary_sheet[1]:  # Accessing the second row directly
    cell.fill = light_green_fill
    cell.font = Font(bold=True)

# Add borders and align cells in summary data
for row in summary_sheet.iter_rows(min_row=2, min_col=1, max_row=summary_sheet.max_row, max_col=2):
    for cell in row:
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="right")

# Define fills for coloring
fill_mismatch = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  # Orange
fill_missing_left = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow
fill_missing_right = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")  # Pink

# Apply color fills in the legend for the "Color" column
for row in summary_sheet.iter_rows(min_row=2, min_col=3, max_row=5, max_col=3):
    cell = row[0]
    if cell.value == "Orange":
        cell.fill = fill_mismatch
    elif cell.value == "Yellow":
        cell.fill = fill_missing_left
    elif cell.value == "Pink":
        cell.fill = fill_missing_right

# Create Comparison Sheet
comparison_sheet = wb.create_sheet(title="Data Comparison")
for r in dataframe_to_rows(comparison_df, index=False, header=True):
    comparison_sheet.append(r)

# Make header bold and enable filter
header_font = Font(bold=True)
# Apply styling to header row in Comparison Sheet
for cell in comparison_sheet[1]:  # Accessing the first row directly
    cell.fill = light_green_fill
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="right", vertical="center")
    cell.border = thin_border
comparison_sheet.auto_filter.ref = comparison_sheet.dimensions  # Apply filter to the entire header row

# Apply conditional formatting
for row in comparison_sheet.iter_rows(min_row=2, min_col=2, max_col=comparison_df.shape[1], max_row=comparison_sheet.max_row):
    for i in range(0, len(columns)*2, 2):
        left_cell = row[i]     # Left column cell
        right_cell = row[i+1]  # Right column cell

        left_value = left_cell.value
        right_value = right_cell.value

        # Add borders
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))
        left_cell.border = thin_border
        right_cell.border = thin_border

        # Check for missing values and mismatches
        if pd.isna(left_value) and not pd.isna(right_value):
            left_cell.fill = fill_missing_left
        elif not pd.isna(left_value) and pd.isna(right_value):
            right_cell.fill = fill_missing_right
        elif left_value != right_value:
            left_cell.fill = fill_mismatch
            right_cell.fill = fill_mismatch

# Check if it's a DataFrame, if not convert it
if isinstance(rows_with_differences, pd.DataFrame):
    rows_to_use = rows_with_differences
else:
    rows_to_use = pd.DataFrame(rows_with_differences)  # Convert list to DataFrame

# Create Difference Sheet
difference_sheet = wb.create_sheet(title="Rows with Differences")
# for r in dataframe_to_rows(pd.DataFrame(rows_with_differences), index=False, header=True):
for r in dataframe_to_rows(rows_to_use, index=False, header=True):
    difference_sheet.append(r)

header_font = Font(bold=True)
for cell in difference_sheet["1:1"]:
# Apply styling to header row in Comparison Sheet
    cell.fill = light_green_fill
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="right", vertical="center")
    cell.border = thin_border
difference_sheet.auto_filter.ref = difference_sheet.dimensions

# for row in difference_sheet.iter_rows(min_row=2, max_row=difference_sheet.max_row, min_col=1, max_col=len(columns)*2):
#     for cell in row:
#         if pd.isna(cell.value):
#             cell.border = thin_border
#             cell.fill = fill_missing_left if cell.column % 2 == 0 else fill_missing_right
#         else:
#             cell.fill = fill_mismatch

for row in difference_sheet.iter_rows(min_row=2, max_row=difference_sheet.max_row, min_col=2, max_col=len(rows_to_use.columns)):
    # for cell in row:
    #     if pd.isna(cell.value):
    #         cell.border = thin_border
    #         cell.fill = fill_missing_left if cell.column % 2 == 0 else fill_missing_right
    #     else:
    #         cell.fill = fill_mismatch

    for i in range(0, len(columns)*2, 2):
        left_cell = row[i]     # Left column cell
        right_cell = row[i+1]  # Right column cell

        left_value = left_cell.value
        right_value = right_cell.value

        # Add borders
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))
        left_cell.border = thin_border
        right_cell.border = thin_border

        # Check for missing values and mismatches
        if pd.isna(left_value) and not pd.isna(right_value):
            left_cell.fill = fill_missing_left
        elif not pd.isna(left_value) and pd.isna(right_value):
            right_cell.fill = fill_missing_right
        elif left_value != right_value:
            left_cell.fill = fill_mismatch
            right_cell.fill = fill_mismatch

# Create No Difference Sheet
no_diff_sheet = wb.create_sheet(title="Rows with No Differences")
for r in dataframe_to_rows(pd.DataFrame(rows_with_no_differences), index=False, header=True):
    no_diff_sheet.append(r)

# Make header bold for no difference sheet
header_font = Font(bold=True)
for cell in no_diff_sheet["1:1"]:
# Apply styling to header row in Comparison Sheet
    cell.fill = light_green_fill
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="right", vertical="center")
    cell.border = thin_border
no_diff_sheet.auto_filter.ref = no_diff_sheet.dimensions

# Adjust column widths
for sheet in wb.worksheets:
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column].width = adjusted_width

# Save the Excel file
wb.save("./financial_data_comparison_side_by_side.xlsx")